import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Выполняем GET запрос к API
url = 'https://weather.sakhalin.gov.ru/api/air/now'

# Добавляем заголовки для имитации браузера
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json',
    'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8',
    'Accept-Encoding': 'gzip, deflate',
    'Connection': 'keep-alive'
}

print("Отправка запроса с User-Agent...")
response = requests.get(url, headers=headers)

# Проверяем статус ответа
print(f"Статус код: {response.status_code}")

if response.status_code == 200:
    # Получаем данные в формате JSON
    data = response.json()
    print(f"Получено записей: {len(data) if isinstance(data, list) else 'не список'}")
    
    # Функция для извлечения нужных полей с русскими названиями
    def process_weather_data(data):
        processed_data = []
        
        for item in data:
            # Создаем базовый словарь с основной информацией на русском
            row = {
                'ID': item.get('id'),
                'Внешний ID': item.get('ext_id'),
                'Населенный пункт': item.get('name'),
                'Широта': item.get('latitude'),
                'Долгота': item.get('longitude'),
                'Дата и время': item.get('date'),
                'Индекс качества воздуха (AQI)': item.get('aqi')
            }
            
            # Извлекаем значения из вложенных словарей с русскими названиями
            if 'temperature' in item and item['temperature']:
                row['Температура (значение)'] = item['temperature'].get('value')
                row['Температура (единица)'] = item['temperature'].get('unit')
                
            if 'pressure' in item and item['pressure']:
                row['Давление (значение)'] = item['pressure'].get('value')
                row['Давление (единица)'] = item['pressure'].get('unit')
                
            if 'humidity' in item and item['humidity']:
                row['Влажность (значение)'] = item['humidity'].get('value')
                row['Влажность (единица)'] = item['humidity'].get('unit')
                
            if 'pm2' in item and item['pm2']:
                row['PM2.5 (значение)'] = item['pm2'].get('value')
                row['PM2.5 (единица)'] = item['pm2'].get('unit')
                
            if 'pm10' in item and item['pm10']:
                row['PM10 (значение)'] = item['pm10'].get('value')
                row['PM10 (единица)'] = item['pm10'].get('unit')
                
            if 'co' in item and item['co']:
                row['Угарный газ CO (значение)'] = item['co'].get('value')
                row['Угарный газ CO (единица)'] = item['co'].get('unit')
                
            if 'no2' in item and item['no2']:
                row['Диоксид азота NO2 (значение)'] = item['no2'].get('value')
                row['Диоксид азота NO2 (единица)'] = item['no2'].get('unit')
                
            if 'so2' in item and item['so2']:
                row['Диоксид серы SO2 (значение)'] = item['so2'].get('value')
                row['Диоксид серы SO2 (единица)'] = item['so2'].get('unit')
                
            if 'o3' in item and item['o3']:
                row['Озон O3 (значение)'] = item['o3'].get('value')
                row['Озон O3 (единица)'] = item['o3'].get('unit')
                
            if 'h2s' in item and item['h2s']:
                row['Сероводород H2S (значение)'] = item['h2s'].get('value')
                row['Сероводород H2S (единица)'] = item['h2s'].get('unit')
                
            if 'wda' in item and item['wda']:
                row['Направление ветра (значение)'] = item['wda'].get('value')
                row['Направление ветра (единица)'] = item['wda'].get('unit')
                
            if 'wva' in item and item['wva']:
                row['Скорость ветра (значение)'] = item['wva'].get('value')
                row['Скорость ветра (единица)'] = item['wva'].get('unit')
                
            if 'ch2o' in item and item['ch2o']:
                row['Формальдегид CH2O (значение)'] = item['ch2o'].get('value')
                row['Формальдегид CH2O (единица)'] = item['ch2o'].get('unit')
            
            processed_data.append(row)
        
        return processed_data
    
    # Обрабатываем данные
    processed_data = process_weather_data(data)
    
    # Создаем DataFrame
    df = pd.DataFrame(processed_data)
    
    # Сохраняем в Excel файл
    filename = f'погодные_данные.xlsx'
    
    # Создаем Excel файл с оформлением
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Погодные данные', index=False)
        
        # Получаем рабочий лист
        worksheet = writer.sheets['Погодные данные']
        
        # Определяем стили
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center')
        cell_alignment_center = Alignment(horizontal='center', vertical='center')
        
        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Применяем стили к заголовкам и настраиваем ширину колонок
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Безопасное вычисление ширины колонки
            try:
                # Получаем данные колонки, заменяем None на пустые строки
                column_data = df.iloc[:, col-1].fillna('').astype(str)
                max_data_length = column_data.map(len).max() if len(column_data) > 0 else 0
                header_length = len(str(df.columns[col-1]))
                max_length = max(max_data_length, header_length)
                adjusted_width = min(max_length + 2, 40)  # Ограничиваем максимальную ширину 40
                worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width
            except Exception as e:
                # Если возникла ошибка, ставим стандартную ширину
                worksheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Применяем стили к ячейкам с данными
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = thin_border
                
                # Выравнивание для разных типов данных
                if col in [1, 2, 4, 5]:  # ID, Внешний ID, Широта, Долгота
                    cell.alignment = cell_alignment_center
                else:
                    cell.alignment = cell_alignment
                
                # Добавляем формат для чисел с плавающей точкой
                if isinstance(cell.value, float):
                    cell.number_format = '0.00'
        
        # Закрепляем первую строку (заголовок)
        worksheet.freeze_panes = 'A2'
        
        # Добавляем фильтр
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Добавляем чередование цветов для строк
        for row in range(2, len(df) + 2):
            if row % 2 == 0:
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Функция для раскраски AQI
    def colorize_aqi():
        wb = load_workbook(filename)
        ws = wb['Погодные данные']
        
        # Находим колонку с AQI
        aqi_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'Индекс качества воздуха (AQI)':
                aqi_col = col
                break
        
        # Раскрашиваем ячейки AQI
        if aqi_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=aqi_col)
                aqi_value = cell.value
                
                if aqi_value is not None:
                    try:
                        aqi_value = float(aqi_value)
                        if aqi_value <= 3:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Зеленый - хороший
                            cell.font = Font(color='006100', bold=True)
                        elif aqi_value <= 4:
                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Желтый - средний
                            cell.font = Font(color='9C6500', bold=True)
                        else:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Красный - плохой
                            cell.font = Font(color='9C0006', bold=True)
                    except (ValueError, TypeError):
                        pass
        
        wb.save(filename)
    
    # Применяем раскраску AQI
    colorize_aqi()
    
    print(f"\n✅ Данные успешно сохранены в файл: {filename}")
    print(f"📊 Всего записей: {len(df)}")
    print(f"📋 Столбцов в файле: {len(df.columns)}")
    print(f"\n🎨 Применено оформление:")
    print(f"  • Синий заголовок с белым текстом")
    print(f"  • Автоподбор ширины колонок")
    print(f"  • Чередование цветов строк")
    print(f"  • Цветовая индикация AQI (зеленый/желтый/красный)")
    print(f"  • Закрепленная шапка")
    print(f"  • Фильтры для данных")
    print(f"\nСписок столбцов:")
    for col in df.columns:
        print(f"  • {col}")
    
else:
    print(f"❌ Ошибка при получении данных: {response.status_code}")
    print(f"Текст ошибки: {response.text}")
